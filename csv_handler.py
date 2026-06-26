"""
csv_handler.py
===============
Converts uploaded CSV or Excel files into a queryable SQLite database.

Supported formats:
  .csv          — plain text CSV (any delimiter)
  .xlsx         — Excel 2007+ (requires openpyxl)
  .xls          — Excel 97-2003 (requires xlrd)

Features:
  - Auto-detects delimiter for CSV (, ; | \t)
  - Auto-detects column data types (INTEGER, REAL, TEXT, DATE)
  - Handles Excel files with multiple sheets — user can pick the sheet
  - Cleans column names (spaces → underscores, removes special chars)
  - Saves as SQLite .db file in uploads/ so it persists across requests
  - Returns a standard conn_info dict compatible with db_connector.py
"""

import os
import re
import csv
import sqlite3
from datetime import datetime, date


# ── Column name cleaner ───────────────────────────────────────────

def clean_column_name(name: str) -> str:
    """Make a column name safe for SQLite."""
    name = str(name).strip()
    name = re.sub(r"[^\w\s]", "", name)
    name = re.sub(r"\s+", "_", name)
    name = name.strip("_")
    if not name:
        name = "column"
    if name[0].isdigit():
        name = "col_" + name
    return name.lower()


def dedupe_headers(raw_headers: list) -> list:
    """Clean and deduplicate column names."""
    seen = {}
    headers = []
    for h in raw_headers:
        clean = clean_column_name(h) or "column"
        if clean in seen:
            seen[clean] += 1
            clean = f"{clean}_{seen[clean]}"
        else:
            seen[clean] = 0
        headers.append(clean)
    return headers


# ── Type inference ────────────────────────────────────────────────

def infer_type(values: list) -> str:
    """
    Infer SQLite column type from sample string values.
    Returns 'INTEGER', 'REAL', or 'TEXT'.
    """
    non_empty = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "TEXT"

    # Try INTEGER
    try:
        for v in non_empty:
            int(str(v).replace(",", ""))
        return "INTEGER"
    except (ValueError, TypeError):
        pass

    # Try REAL
    try:
        for v in non_empty:
            float(str(v).replace(",", ""))
        return "REAL"
    except (ValueError, TypeError):
        pass

    return "TEXT"


def coerce_value(value, dtype: str):
    """Cast a value to the inferred Python type."""
    # Handle Excel date/datetime objects directly
    if isinstance(value, (datetime, date)):
        return str(value)

    if value is None:
        return None

    value_str = str(value).strip()
    if value_str == "":
        return None

    if dtype == "INTEGER":
        try:
            return int(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return None
    if dtype == "REAL":
        try:
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return None
    return value_str


# ── CSV reading ───────────────────────────────────────────────────

def detect_delimiter(file_path: str) -> str:
    """Sniff the delimiter from the first few KB of the file."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def read_csv_file(file_path: str, delimiter: str = None) -> tuple[list, list]:
    """
    Read a CSV file and return (headers_raw, rows).
    rows is a list of lists of strings.
    """
    # Try UTF-8 first, then fall back to latin-1 for Windows-exported CSVs
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(file_path, "r", encoding=encoding, errors="strict") as f:
                if delimiter is None:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                        used_delim = dialect.delimiter
                    except csv.Error:
                        used_delim = ","
                else:
                    used_delim = delimiter

                reader = csv.reader(f, delimiter=used_delim)
                try:
                    raw_headers = next(reader)
                except StopIteration:
                    raise ValueError("CSV file is empty.")

                if not raw_headers or all(h.strip() == "" for h in raw_headers):
                    raise ValueError("CSV file has no headers in the first row.")

                rows = list(reader)

            return raw_headers, rows, used_delim

        except (UnicodeDecodeError, UnicodeError):
            continue

    raise ValueError("Could not decode the CSV file. Try saving it as UTF-8 from Excel (File → Save As → CSV UTF-8).")


# ── Excel reading ─────────────────────────────────────────────────

def get_excel_sheets(file_path: str) -> list:
    """Return list of sheet names in an Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to read .xlsx files.\n"
                "Run: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets

    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required to read .xls files.\n"
                "Run: pip install xlrd"
            )
        wb = xlrd.open_workbook(file_path)
        return wb.sheet_names()

    raise ValueError(f"Not an Excel file: {ext}")


def read_excel_file(file_path: str, sheet_name: str = None) -> tuple[list, list]:
    """
    Read an Excel file and return (headers_raw, rows).
    rows is a list of lists (values may be Python native types).
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        all_rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                all_rows.append(list(row))
        wb.close()

        if not all_rows:
            raise ValueError("Excel sheet is empty.")

        raw_headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(all_rows[0])]
        rows = all_rows[1:]
        return raw_headers, rows

    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise ImportError("xlrd is required. Run: pip install xlrd")

        wb = xlrd.open_workbook(file_path)

        if sheet_name:
            try:
                ws = wb.sheet_by_name(sheet_name)
            except xlrd.XLRDError:
                ws = wb.sheet_by_index(0)
        else:
            ws = wb.sheet_by_index(0)

        if ws.nrows == 0:
            raise ValueError("Excel sheet is empty.")

        raw_headers = [str(ws.cell_value(0, c)) if ws.cell_value(0, c) != "" else f"col{c}"
                       for c in range(ws.ncols)]

        rows = []
        for r in range(1, ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                # Convert xlrd date serial numbers to strings
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(str(dt.date()))
                    except Exception:
                        row.append(str(cell.value))
                else:
                    row.append(cell.value)
            rows.append(row)

        return raw_headers, rows

    raise ValueError(f"Unsupported Excel format: {ext}")


# ── Shared: write rows to SQLite ──────────────────────────────────

def rows_to_sqlite(
    raw_headers: list,
    rows: list,
    table_name: str,
    db_path: str,
) -> dict:
    """
    Write rows into a SQLite table and return a preview dict.
    raw_headers : list of original header strings
    rows        : list of lists (values can be any Python type)
    """
    headers   = dedupe_headers(raw_headers)
    n_cols    = len(headers)

    # Infer types from first 200 data rows
    sample = rows[:200]
    col_types = []
    for i in range(n_cols):
        col_values = [r[i] if i < len(r) else None for r in sample]
        col_types.append(infer_type(col_values))

    # Write to SQLite
    conn = sqlite3.connect(db_path)
    cur  = conn.cursor()
    cur.execute(f"DROP TABLE IF EXISTS [{table_name}]")

    col_defs = ", ".join(f"[{h}] {t}" for h, t in zip(headers, col_types))
    cur.execute(f"CREATE TABLE [{table_name}] ({col_defs})")

    placeholders = ", ".join("?" * n_cols)
    inserted = 0
    errors   = 0

    for row in rows:
        padded = (list(row) + [None] * n_cols)[:n_cols]
        values = [coerce_value(padded[i], col_types[i]) for i in range(n_cols)]
        try:
            cur.execute(f"INSERT INTO [{table_name}] VALUES ({placeholders})", values)
            inserted += 1
        except Exception:
            errors += 1

    conn.commit()
    conn.close()

    # Get preview rows
    conn2 = sqlite3.connect(db_path)
    cur2  = conn2.cursor()
    cur2.execute(f"SELECT * FROM [{table_name}] LIMIT 5")
    preview_rows = [list(r) for r in cur2.fetchall()]
    conn2.close()

    return {
        "table_name":     table_name,
        "columns":        headers,
        "col_types":      col_types,
        "total_rows":     inserted,
        "skipped_rows":   errors,
        "preview_rows":   preview_rows,
        "schema_display": f"{table_name}({', '.join(headers)})",
        "db_path":        db_path,
    }


# ── Public API ────────────────────────────────────────────────────

def is_excel(filename: str) -> bool:
    return filename.lower().endswith((".xlsx", ".xls"))

def is_csv(filename: str) -> bool:
    return filename.lower().endswith(".csv")

def is_supported(filename: str) -> bool:
    return is_csv(filename) or is_excel(filename)


def file_to_sqlite(
    file_path: str,
    table_name: str = None,
    output_dir: str = "uploads",
    delimiter: str = None,
    sheet_name: str = None,
) -> tuple[dict, dict]:
    """
    Convert a CSV or Excel file into a SQLite database.

    Parameters
    ----------
    file_path  : full path to the uploaded file
    table_name : SQL table name (defaults to filename stem)
    output_dir : where to save the .db file
    delimiter  : CSV only — override auto-detection
    sheet_name : Excel only — which sheet to read (defaults to first/active)

    Returns
    -------
    conn_info  : dict for db_connector
    preview    : dict with schema info and sample rows
    """
    os.makedirs(output_dir, exist_ok=True)

    filename = os.path.basename(file_path)
    stem     = os.path.splitext(filename)[0]

    if not table_name:
        table_name = clean_column_name(stem) or "data"

    db_filename = stem + "_data.db"
    db_path     = os.path.join(os.path.abspath(output_dir), db_filename)

    # Read file
    if is_excel(filename):
        raw_headers, rows = read_excel_file(file_path, sheet_name=sheet_name)
        used_delimiter    = None
    elif is_csv(filename):
        raw_headers, rows, used_delimiter = read_csv_file(file_path, delimiter=delimiter)
    else:
        raise ValueError(f"Unsupported file type. Please upload .csv, .xlsx, or .xls")

    if not rows:
        raise ValueError("File has headers but no data rows.")

    preview = rows_to_sqlite(raw_headers, rows, table_name, db_path)
    if used_delimiter:
        preview["delimiter"] = used_delimiter

    label = f"{filename} → {table_name}"
    if is_excel(filename) and sheet_name:
        label = f"{filename} [{sheet_name}] → {table_name}"

    conn_info = {
        "type":   "sqlite",
        "path":   db_path,
        "label":  label,
        "source": "upload",
    }

    return conn_info, preview


def get_file_info(file_path: str, delimiter: str = None) -> dict:
    """
    Quick scan — returns headers, inferred types, 3 sample rows, and sheet list.
    Used by the /db/preview-csv endpoint before actually loading.
    """
    filename = os.path.basename(file_path)

    try:
        if is_excel(filename):
            sheets        = get_excel_sheets(file_path)
            raw_headers, rows = read_excel_file(file_path)
            used_delimiter    = None
        elif is_csv(filename):
            sheets             = []
            raw_headers, rows, used_delimiter = read_csv_file(file_path, delimiter=delimiter)
        else:
            return {"error": f"Unsupported file type. Upload .csv, .xlsx, or .xls"}

    except ImportError as e:
        return {"error": str(e)}
    except Exception as e:
        return {"error": str(e)}

    headers   = dedupe_headers(raw_headers)
    sample    = rows[:5]
    col_types = []
    for i in range(len(headers)):
        vals = [r[i] if i < len(r) else None for r in sample]
        col_types.append(infer_type(vals))

    sample_rows = [[str(v) if v is not None else "" for v in (r[:len(headers)])] for r in rows[:3]]

    return {
        "raw_headers":  [str(h) for h in raw_headers],
        "headers":      headers,
        "col_types":    col_types,
        "delimiter":    used_delimiter,
        "sample_rows":  sample_rows,
        "sheets":       sheets,
        "file_type":    "excel" if is_excel(filename) else "csv",
        "success":      True,
    }


# ── Legacy aliases (keep old imports working) ─────────────────────
def csv_to_sqlite(csv_path, table_name=None, output_dir="uploads", delimiter=None):
    return file_to_sqlite(csv_path, table_name=table_name,
                          output_dir=output_dir, delimiter=delimiter)

def get_csv_info(csv_path, delimiter=None):
    return get_file_info(csv_path, delimiter=delimiter)
